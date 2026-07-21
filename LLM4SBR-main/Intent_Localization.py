import pandas as pd
from sklearn.metrics.pairwise import cosine_similarity
import torch
import os
from tqdm import tqdm
from transformers import BertTokenizer, BertModel
import ast
import torch.nn.functional as F
import torch.nn as nn
import openpyxl


# 意图定位模块：
# 先用本地 BERT 把 LLM 生成的兴趣文本和候选 item 名称都编码成向量，
# 再用余弦相似度找到最接近的 top-5 item 名称，
# 最后用相似度加权的 item 名称 embedding 作为“定位后”的语义兴趣 embedding。
# 这样做的目的是把 LLM 的自由文本回答约束回数据集里的真实 item 空间，降低幻觉影响。

device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Using device: {device}")

# 从本地目录加载 BERT；如果该目录不存在，需要先下载/放置对应模型文件。
tokenizer = BertTokenizer.from_pretrained('../../LLMs/bert-base-uncased')
model = BertModel.from_pretrained("../../LLMs/bert-base-uncased").to(device)

def text_to_embedding(text):
    """把一段文本编码为 BERT 句向量。

    长文本按字符长度切块，分别编码后拼接；短文本直接取 last_hidden_state 的均值。
    """
    max_chunk_len = 200
    if len(text) > max_chunk_len:
        # 避免过长文本超过 BERT 输入长度，这里按 200 字符切分。
        text_chunks = [text[i:i + max_chunk_len] for i in range(0, len(text), max_chunk_len)]
        chunks_embeddings = []
        for chunk in text_chunks:
            inputs = tokenizer.encode(chunk, return_tensors="pt").to(device)  # Move inputs to GPU
            with torch.no_grad():
                outputs = model(inputs)
            # 对 token 维度取平均，得到该 chunk 的句向量。
            embeddings = outputs.last_hidden_state.mean(dim=1)
            chunks_embeddings.append(embeddings)
        # 注意：这里是按特征维拼接，长文本会得到超过 768 维的向量。
        # 后续 unify_second_dimension 会再线性映射回 768 维。
        final_encoding = torch.cat(chunks_embeddings, dim=1)
        return final_encoding
    else:
        tokens = tokenizer(text, return_tensors='pt')
        tokens = {key: val.to(device) for key, val in tokens.items()}  # Move all token tensors to GPU
        with torch.no_grad():
            outputs = model(**tokens)
        embeddings = outputs.last_hidden_state.mean(dim=1)
        return embeddings

def unify_second_dimension(embedding):
    """把任意第二维长度的 embedding 统一到 768 维。

    对超过 768 维的长文本拼接向量，使用一个临时 Linear 层投影。
    """
    target_dim = 768
    if embedding.size(1) != target_dim:
        linear_layer = torch.nn.Linear(embedding.size(1), target_dim).to(device)  # Move linear layer to GPU
        processed_embedding = linear_layer(embedding)
    else:
        processed_embedding = embedding.clone()
    return processed_embedding

def main():
    """读取 LLM 输出与候选 item 名称，生成定位后的 session embedding 表。"""
    # type 控制处理训练集还是测试集，如 tra/tes；当前脚本写死为 tra。
    type = 'tra'
    xlsx_file_path = "Search_data/ml-1m/{}_session_long.xlsx".format(type)
    workbook = openpyxl.load_workbook(xlsx_file_path)
    sheet = workbook.active
    keys = []
    values = []

    key_column_index = 1
    value_column_index = 2

    # 第一列通常是 prompt，第二列是 LLM 的回答文本。
    for row in sheet.iter_rows(min_row=2, values_only=True):
        keys.append(row[key_column_index - 1])
        values.append(row[value_column_index - 1])

    workbook.close()

    df = pd.DataFrame({'Key': keys, 'Value': values})

    # movie_name.xlsx/name.xlsx 这类文件保存数据集中真实 item 名称，作为定位候选库。
    xlsx_file_path2 = "Search_data/ml-1m/movie_name.xlsx"
    workbook_1 = openpyxl.load_workbook(xlsx_file_path2)
    sheet_1 = workbook_1.active
    names = []
    names_column_index = 1
    for row in sheet_1.iter_rows(min_row=2, values_only=True):
        names.append(row[names_column_index - 1])
    workbook_1.close()

    knowledge_df = pd.DataFrame({'name': names})
    # 分别编码 LLM 回答文本和候选 item 名称。
    df['value_emb'] = df['Value'].apply(lambda x: text_to_embedding(str(x)))
    knowledge_df['name_emb'] = knowledge_df['name'].apply(lambda x: text_to_embedding(str(x)))
    # 保证后续相似度计算都是 768 维。
    df['value_emb'] = df['value_emb'].apply(lambda x: unify_second_dimension(x))
    knowledge_df['name_emb'] = knowledge_df['name_emb'].apply(lambda x: unify_second_dimension(x))

    print(df.head(5))
    print(knowledge_df.head(5))

    for index, row in tqdm(df.iterrows(), total=len(df), desc="Processing Rows", unit="row"):
        value_embedding = row['value_emb']
        # 把所有候选 item 名称向量堆叠成矩阵，逐条与当前 LLM 回答向量计算余弦相似度。
        name_embeddings = torch.stack(knowledge_df['name_emb'].tolist()).view(len(knowledge_df), -1).to(device)
        similarities = cosine_similarity(value_embedding.cpu().detach().numpy(), name_embeddings.cpu().detach().numpy())
        # top5_names 是最接近 LLM 回答的真实 item 名称，用于人工查看定位效果。
        top5_indices = torch.topk(torch.tensor(similarities.flatten()), 5)[1]
        top5_name = knowledge_df['name'].iloc[top5_indices].tolist()
        df.at[index, 'top5_names'] = str(top5_name)
        top5_name_embs = knowledge_df['name_emb'].iloc[top5_indices].tolist()
        similarities_tensor = torch.tensor(similarities.flatten(), device=device)
        # 用相似度加权 top-5 候选 item 的 embedding，得到最终 long_emb。
        # 注意：这里 zip(top5_name_embs, similarities_tensor) 会拿 top5 embedding
        # 与 similarities_tensor 的前 5 个值相乘；若想严格使用 top5 相似度，应取 top5_indices 对应值。
        updated_value_emb = sum([emb * sim for emb, sim in zip(top5_name_embs, similarities_tensor)]) / 5
        df.at[index, 'long_emb'] = str(updated_value_emb.cpu().tolist())  # Move to CPU for storage

    print(df.head(5))
    # 输出文件会被 main.py 读取，用作模型训练/测试时的 LLM 长期兴趣 embedding。
    df.to_excel('Search_data/ml-1m/{}_session_long_emb.xlsx'.format(type))

if __name__ == '__main__':
    main()
